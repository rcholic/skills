#!/usr/bin/env python3
"""
解析 TRTC 场景配置 Excel 和巡检结果 Excel，输出结构化的 JSON 数据。

用法:
  python3 parse_excel.py <excel_file_path> [--type config|inspect]

  --type config   : 解析场景配置模板 Excel（如 "场景配置_视频类-秀场直播_*.xlsx"）
  --type inspect  : 解析巡检结果 Excel（如 "巡检结果_*.xlsx"）
  不传 --type     : 自动根据文件名猜测类型
"""

import json
import sys
import os

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl


def safe_str(cell_value):
    """安全地将单元格值转为字符串。"""
    if cell_value is None:
        return ""
    return str(cell_value).strip()


def parse_config_excel(filepath):
    """解析场景配置模板 Excel，返回结构化数据。"""
    if not os.path.isfile(filepath):
        raise FileNotFoundError(f"文件不存在: {filepath}")

    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]

    result = {
        "type": "scene_config",
        "file": os.path.basename(filepath),
        "basic_info": {},
        "sections": {}
    }

    current_section = None
    rows = list(ws.iter_rows(values_only=True))

    # 已知的基础信息字段
    basic_fields = {"场景名称", "应用场景", "产品类型", "类似客户", "创建人", "更新人", "创建时间", "更新时间"}
    # 已知的分区标题
    section_titles = {"房间管理", "音频", "视频参数", "3A"}
    # 已知的标题行/跳过行
    skip_headers = {"场景配置详情", "基础信息"}

    i = 0
    while i < len(rows):
        cells = [safe_str(c) for c in rows[i]]

        # 跳过空行
        if not any(cells):
            i += 1
            continue

        # 跳过标题行
        if cells[0] in skip_headers:
            i += 1
            continue

        # 识别基础信息字段
        if cells[0] in basic_fields:
            result["basic_info"][cells[0]] = cells[1]
            i += 1
            continue

        # 识别分区标题
        if cells[0] in section_titles and (len(cells) < 2 or cells[1] == ""):
            current_section = cells[0]
            result["sections"][current_section] = []
            i += 1
            continue

        # 跳过表头行
        if cells[0] == "配置项" and cells[1] == "参数描述":
            i += 1
            continue

        # 解析配置项
        if current_section and cells[0]:
            config_item = {
                "config_name": cells[0],
                "description": cells[1],
                "target_value": cells[2],
                "code_example": cells[3] if len(cells) > 3 else ""
            }
            result["sections"][current_section].append(config_item)

        i += 1

    return result


def parse_inspect_excel(filepath):
    """解析巡检结果 Excel，返回结构化数据。"""
    if not os.path.isfile(filepath):
        raise FileNotFoundError(f"文件不存在: {filepath}")

    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]

    result = {
        "type": "inspection_result",
        "file": os.path.basename(filepath),
        "basic_info": {},
        "sections": {},
        "sdk_versions": []
    }

    current_section = None
    rows = list(ws.iter_rows(values_only=True))

    # 已知的基础信息字段
    basic_fields = {"任务ID", "SDK AppID", "客户名称", "创建者", "巡检时间范围"}
    # 已知的分区标题
    section_titles = {"音频", "房间管理", "视频参数", "3A", "SDK版本"}

    i = 0
    while i < len(rows):
        cells = [safe_str(c) for c in rows[i]]

        if not any(cells):
            i += 1
            continue

        # 基础信息
        if cells[0] in basic_fields:
            result["basic_info"][cells[0]] = cells[1]
            i += 1
            continue

        if cells[0] == "基础信息":
            i += 1
            continue

        # 分区
        if cells[0] in section_titles and (len(cells) < 2 or cells[1] == ""):
            current_section = cells[0]
            if current_section != "SDK版本":
                result["sections"][current_section] = []
            i += 1
            continue

        # 跳过表头
        if cells[0] == "配置项" and cells[1] in ("参数描述", ""):
            i += 1
            continue

        if cells[0] == "平台" and cells[1] == "版本":
            i += 1
            continue

        # SDK版本
        if current_section == "SDK版本" and cells[0]:
            result["sdk_versions"].append({
                "platform": cells[0],
                "version": cells[1],
                "count": cells[2]
            })
            i += 1
            continue

        # 配置项
        if current_section and current_section != "SDK版本" and cells[0]:
            item = {
                "config_name": cells[0],
                "description": cells[1],
                "platform": cells[2],
                "current_value": cells[3],
                "code_example": cells[4] if len(cells) > 4 else ""
            }
            result["sections"][current_section].append(item)

        i += 1

    return result


def guess_type(filepath):
    """根据文件名猜测 Excel 类型。"""
    basename = os.path.basename(filepath)
    if "场景配置" in basename or "config" in basename.lower():
        return "config"
    elif "巡检结果" in basename or "inspect" in basename.lower():
        return "inspect"
    else:
        return None


def main():
    if len(sys.argv) < 2:
        print("用法: python3 parse_excel.py <excel_file_path> [--type config|inspect]")
        sys.exit(1)

    filepath = sys.argv[1]
    file_type = None

    if "--type" in sys.argv:
        idx = sys.argv.index("--type")
        if idx + 1 < len(sys.argv):
            file_type = sys.argv[idx + 1]

    if not file_type:
        file_type = guess_type(filepath)

    if not file_type:
        print(f"无法识别文件类型: {filepath}，请使用 --type config 或 --type inspect 指定",
              file=sys.stderr)
        sys.exit(1)

    try:
        if file_type == "config":
            result = parse_config_excel(filepath)
        else:
            result = parse_inspect_excel(filepath)

        print(json.dumps(result, ensure_ascii=False, indent=2))
    except FileNotFoundError as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(json.dumps({"error": f"解析失败: {str(e)}"}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
